from PIL import Image
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Helper function: RGB -> hexadecimal conversion
def rgb_to_hex(r, g, b):
    return f"{r:02X}{g:02X}{b:02X}"


image = Image.open("cat-small-jpeg.jpg")
white = 255
reds = []
blues = []
greens = []

# Extract RGB values from each pixel
for x in range(1, 64):
    for y in range(1, 63):
        pixel = image.getpixel((x, y))
        r, g, b = pixel[:3]
        # If the pixel is non-white, add the co-ordinate and RBG-value to the array
        if (r != white): reds.append((x, y, r))
        if (g != white): greens.append((x, y, g))
        if (b != white): blues.append((x, y, b))
# Create dataframes from
rdf = pd.DataFrame(reds, columns=["X", "Y", "R"])
gdf = pd.DataFrame(greens, columns=[["X", "Y", "G"]])
bdf = pd.DataFrame(blues, columns=["X", "Y", "B"])

# Test that the dataframes are working
print(rdf.head())
print(gdf.head())
print(bdf.head())
# Export dataframes to Excel
rdf.to_excel("rgb-table-r.xlsx", index=False, engine="openpyxl")
gdf.to_excel("rgb-table-g.xlsx", index=False, engine="openpyxl")
bdf.to_excel("rgb-table-b.xlsx", index=False, engine="openpyxl")
print("Excel tables printed")

# Populate a second Excel sheet with the RGB val in cell (x, y) (Red)
wb = load_workbook("rgb-graph-r.xlsx")
sheet = wb.active
for index, row in rdf.iterrows():
    x = row["X"]
    y = row["Y"]
    r = row["R"]
    sheet.cell(row=y, column=x, value=r) # Add val to cell
    # Add colour shading to cell
    hex_colour = rgb_to_hex(r, 0, 0)
    fill_colour = PatternFill(start_color=hex_colour, end_color=hex_colour, fill_type="solid")
    sheet.cell(row=y, column=x).fill = fill_colour
wb.save("rgb-graph-r.xlsx")

# Green pixel graph
wb = load_workbook("rgb-graph-g.xlsx")
sheet = wb.active
for index, row in gdf.iterrows():
    x = row["X"]
    y = row["Y"]
    g = row["G"]
    sheet.cell(row=y, column=x, value=g) # Add val to cell
    # Add colour shading to cell
    hex_colour = rgb_to_hex(0, g, 0)
    fill_colour = PatternFill(start_color=hex_colour, end_color=hex_colour, fill_type="solid")
    sheet.cell(row=y, column=x).fill = fill_colour
wb.save("rgb-graph-g.xlsx")

# Blue pixel graph
wb = load_workbook("rgb-graph-b.xlsx")
sheet = wb.active
for index, row in bdf.iterrows():
    x = row["X"]
    y = row["Y"]
    b = row["B"]
    sheet.cell(row=y, column=x, value=b) # Add val to cell
    # Add colour shading to cell
    hex_colour = rgb_to_hex(0, 0, b)
    fill_colour = PatternFill(start_color=hex_colour, end_color=hex_colour, fill_type="solid")
    sheet.cell(row=y, column=x).fill = fill_colour
wb.save("rgb-graph-b.xlsx")

print("Graphs created with shading")